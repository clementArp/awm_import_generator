import json
import sqlite3
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

# ============================================================
# Constantes / Config
# ============================================================

# Excel tables & columns
TABLE_MOTOR_PREFIX = "T_Mot"
COL_MOTOR_AXNAME = "Repère"
COL_MOTOR_GEAR = "Réducteur"
COL_MOTOR_FEED_CST = "Feed constant"
COL_MOTOR_TYPE = "Type"
MOTOR_TYPE_TO_KEEP = "MB"
MOTOR_PREFIX_TO_ADD = "V"

# CSV columns
COL_CSV_AXNAME = "axname"
COL_CSV_GEAR = "refGearBox"
COL_CSV_FEED_CONSTANT = "feedconstant"

# ============================================================
# Création dossier de sortie
# ============================================================

OUT_DIR = Path("out")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# Utilitaires Excel
# ============================================================

def _normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def table_to_list(ws, table_name: str, wanted_columns: Optional[Iterable[str]] = None) -> List[Dict[str, Any]]:
    """Convertit une table Excel (openpyxl) en liste de dicts."""
    if table_name not in ws.tables:
        return []

    table = ws.tables[table_name]
    rows = list(ws[table.ref])
    if not rows:
        return []

    headers = [_normalize_header(c.value) for c in rows[0]]
    out: List[Dict[str, Any]] = []

    for row in rows[1:]:
        values = [c.value for c in row]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        d = {}
        for i, h in enumerate(headers):
            if wanted_columns is None or h in wanted_columns:
                d[h] = values[i]
        out.append(d)

    return out

def read_excel(excel_path: Path) -> Dict[str, Any]:
    """Lit toutes les feuilles et récupère : moteurs"""
    wb = load_workbook(excel_path, data_only=True)

    data = {
        "motors": [],
    }

    for ws in wb.worksheets:
        tables_in_sheet = list(ws.tables.keys())

        # Moteurs
        for table_name in [t for t in tables_in_sheet if t.startswith(TABLE_MOTOR_PREFIX)]:
            data["motors"].extend(table_to_list(ws, table_name))

    return data

# ============================================================
# I/O Console (ask_*)
# ============================================================

def ask_path(prompt: str, allowed_suffixes: Tuple[str, ...]) -> Optional[Path]:
    while True:
        path_str = input(prompt).strip().strip('"')
        if not path_str:
            print("Aucun fichier sélectionné.")
            return None

        try:
            path = Path(path_str).resolve()
        except OSError as e:
            print("Chemin invalide :", e)
            continue

        if not path.exists():
            print("Le fichier n'existe pas.")
            continue

        if not path.is_file():
            print("Le chemin indiqué n'est pas un fichier.")
            continue

        if path.suffix.lower() not in allowed_suffixes:
            print(f"Veuillez sélectionner un fichier parmi : {', '.join(allowed_suffixes)}")
            continue

        return path


def ask_excel_file() -> Optional[Path]:
    return ask_path("Chemin du fichier Excel : ", (".xlsx", ".xlsm", ".xls"))

# ============================================================
# Exports CSV (file to import in AWM)
# ============================================================

def _csv_line(*values: Any) -> str:
    return ";".join(str(v) for v in values) + ";"

def export_motors_csv(motors: List[Dict[str, Any]], out_path: Path) -> None:
    lines: List[str] = []

    lines.append(_csv_line(COL_CSV_AXNAME, COL_CSV_GEAR, COL_CSV_FEED_CONSTANT))

    for m in motors:
        mtype = m.get(COL_MOTOR_TYPE)
        if mtype is None:
            continue
        if MOTOR_TYPE_TO_KEEP.lower() not in str(mtype).lower():
            continue
        name = m.get(COL_MOTOR_AXNAME)
        if name is None:
            continue
        axname = MOTOR_PREFIX_TO_ADD + str(name)
        gear = m.get(COL_MOTOR_GEAR)
        if gear is None:
            print(f"Réducteur manquant pour {axname} => non généré.")
            continue
        feed_cst = m.get(COL_MOTOR_FEED_CST)
        try:
            feed_cst_float = float(feed_cst)
        except (ValueError, TypeError):
            print(f"Feed constant invalide pour {axname} : {feed_cst} => non généré.")
            continue

        lines.append(_csv_line(axname, gear, feed_cst_float))

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

# ============================================================
# Main
# ============================================================

def main() -> None:
    excel_path = ask_excel_file()
    if not excel_path:
        return
    
    print("Lecture du fichier Excel...")
    data = read_excel(excel_path)

    # Exports CSV
    print("Export des CSV...")
    export_motors_csv(data["motors"], OUT_DIR / "motor.csv")


if __name__ == "__main__":
    main()


