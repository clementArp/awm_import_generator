import json
import sqlite3
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


# ============================================================
# Constantes / Config
# ============================================================

BASE_ID_FAULT_DESCRIPTION = 1_00_00_000  # [1_00_00_000; 2_00_00_000[
BASE_ID_BUTTON_TEXT = 2_00_01_000  # [2_00_01_000; 2_00_02_000[
BASE_ID_BUTTON_DESCRIPTION = 2_00_02_000  # [2_00_02_000; 2_00_03_000[
BASE_ID_BYPASS_TEXT = 2_00_03_000  # [2_00_03_000; 2_00_04_000[
BASE_ID_BYPASS_DESCRIPTION = 2_00_04_000  # [2_00_04_000; 2_00_05_000[

LANGUAGE_ARP = "arp"
LANGUAGE_FR = "fr"
LANGUAGE_EN = "en"
LANGUAGE_ES = "es"
LANGUAGE_DE = "de"
LANGUAGE_GR = "gr"

TRANSLATE = {
    "defaut": {LANGUAGE_ARP: 0, LANGUAGE_FR: 0, LANGUAGE_EN: 1, LANGUAGE_ES: 2, LANGUAGE_DE: 3, LANGUAGE_GR: 4},
    "bdd": {LANGUAGE_ARP: 0, LANGUAGE_FR: 1, LANGUAGE_EN: 2, LANGUAGE_ES: 3, LANGUAGE_DE: 4, LANGUAGE_GR: 5},
}

# Excel tables & columns
CELL_EM_PREFIX = "B3"

TABLE_SOMMAIRE = "T_Sommaire"
COL_SOMMAIRE_MODULE = "N° Module"
COL_SOMMAIRE_NUM_MACHINE = "N° Machine"
COL_SOMMAIRE_NUM_MODULE = "N° Unit"
COL_SOMMAIRE_NOM_LANGUE_1 = "Nom Langue 1"
COL_SOMMAIRE_NOM_LANGUE_2 = "Nom Langue 2"

TABLE_DEFAULT_PREFIX = "T_Defaut"
COL_DEFAUT_NUM = "Code défaut"
COL_DEFAUT_RESOLUTION_ARP = "Résolution ARP"
COL_DEFAUT_RESOLUTION_CLIENT = "Résolution Client"

TABLE_BYPASS = "T_RecapShunt"
TABLE_BYPASS_EM_PREFIX = "T_Shunt_U"
COL_BYPASS_NUM = "N°"
COL_BYPASS_NUM_MODULE = "N° Module"
COL_BYPASS_DESIGNATION_ARP = "Désignation ARP"
COL_BYPASS_DESIGNATION_CLIENT = "Désignation Client"
COL_BYPASS_DESCRIPTION_ARP = "Description ARP"
COL_BYPASS_DESCRIPTION_CLIENT = "Description Client"
COL_BYPASS_ALIAS = "Repère"
COL_BYPASS_ALIAS_EM = "Shunt"
COL_BYPASS_ALIAS_EM_IN_EM = "Repère"
COL_BYPASS_CHECK = "Check1"
check_bypass_is_ok = lambda b: str(b.get(COL_BYPASS_CHECK, "")).strip().lower() in (1, "1", True, "true")

TABLE_BUTTON = "T_RecapBtn"
TABLE_BUTTON_EM_PREFIX = "T_Action_U"
COL_BUTTON_NUM = "N°"
COL_BUTTON_NUM_MODULE = "N° Module"
COL_BUTTON_DESIGNATION_ARP = "Désignation ARP"
COL_BUTTON_DESIGNATION_CLIENT = "Désignation Client"
COL_BUTTON_DESCRIPTION_ARP = "Description ARP"
COL_BUTTON_DESCRIPTION_CLIENT = "Description Client"
COL_BUTTON_ALIAS = "Repère"
COL_BUTTON_ALIAS_EM = "Btn"
COL_BUTTON_ALIAS_EM_IN_EM = "Repère"
COL_BUTTON_CHECK = "Check1"
check_button_is_ok = lambda b: str(b.get(COL_BUTTON_CHECK, "")).strip().lower() in (1, "1", True, "true")

# JSON keys
JSON_BYPASS_NUM = "num"
JSON_BYPASS_NUM_MACHINE = "num_machine"
JSON_BYPASS_NUM_MODULE = "num_em"
JSON_BYPASS_ALIAS = "alias"

JSON_BUTTON_NUM = "num"
JSON_BUTTON_NUM_MACHINE = "num_machine"
JSON_BUTTON_NUM_MODULE = "num_em"
JSON_BUTTON_ALIAS = "alias"

# ============================================================
# Création dossier de sortie
# ============================================================

OUT_DIR = Path("out")
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# Utilitaires Excel
# ============================================================


def _normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def table_to_list(ws, table_name: str, wanted_columns: Optional[Iterable[str]] = None) -> List[Dict[str, Any]]:
    """Convertit une table Excel (openpyxl) en liste de dicts."""
    if table_name not in ws.tables:
        return []

    table = ws.tables[table_name]
    rows = list(ws[table.ref])
    if not rows:
        return []

    headers = [_normalize_header(c.value) for c in rows[0]]
    out: List[Dict[str, Any]] = []

    for row in rows[1:]:
        values = [c.value for c in row]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        d = {}
        for i, h in enumerate(headers):
            if wanted_columns is None or h in wanted_columns:
                d[h] = values[i]
        out.append(d)

    return out


def read_excel(excel_path: Path) -> Dict[str, Any]:
    """Lit toutes les feuilles et récupère : defauts, bypass, buttons, modules_cfg."""
    wb = load_workbook(excel_path, data_only=True)
    data = {
        "defauts": [],
        "bypass": [],
        "buttons": [],
        "bypass_em": {},
        "buttons_em": {},
        "modules_cfg": {},  # module -> cfg
    }

    for ws in wb.worksheets:

        sheet_name = ws.title

        sheet_em = ws[CELL_EM_PREFIX].value

        tables_in_sheet = list(ws.tables.keys())

        # Sommaire -> modules_cfg
        if TABLE_SOMMAIRE in tables_in_sheet:
            for item in table_to_list(ws, TABLE_SOMMAIRE):
                module = item.get(COL_SOMMAIRE_MODULE)
                if module is None:
                    continue
                try:
                    num_machine = int(item.get(COL_SOMMAIRE_NUM_MACHINE))
                    num_module = int(item.get(COL_SOMMAIRE_NUM_MODULE))
                except (TypeError, ValueError):
                    print(f"Erreur conversion num_machine/num_module pour module {module}")
                    continue

                data["modules_cfg"][module] = {
                    "num_machine": num_machine,
                    "num_module": num_module,
                    "nom_langue_1": item.get(COL_SOMMAIRE_NOM_LANGUE_1, "") or "",
                    "nom_langue_2": item.get(COL_SOMMAIRE_NOM_LANGUE_2, "") or "",
                }

        # Défauts
        for table_name in [t for t in tables_in_sheet if t.startswith(TABLE_DEFAULT_PREFIX)]:
            data["defauts"].extend(table_to_list(ws, table_name))

        # Bypass / Buttons
        if TABLE_BYPASS in tables_in_sheet:
            data["bypass"].extend(table_to_list(ws, TABLE_BYPASS))

        if TABLE_BUTTON in tables_in_sheet:
            data["buttons"].extend(table_to_list(ws, TABLE_BUTTON))

        data["bypass_em"][sheet_em] = {}
        for table_name in [t for t in tables_in_sheet if t.startswith(TABLE_BYPASS_EM_PREFIX)]:
            rows = table_to_list(ws, table_name)
            for row in rows:
                if not check_bypass_is_ok(row):
                    continue
                if not COL_BYPASS_ALIAS_EM_IN_EM in row:
                    continue
                data["bypass_em"][sheet_em][row[COL_BYPASS_ALIAS_EM_IN_EM]] = row

        data["buttons_em"][sheet_em] = {}
        for table_name in [t for t in tables_in_sheet if t.startswith(TABLE_BUTTON_EM_PREFIX)]:
            rows = table_to_list(ws, table_name)
            for row in rows:
                if not check_button_is_ok(row):
                    continue
                if not COL_BUTTON_ALIAS_EM_IN_EM in row:
                    continue
                data["buttons_em"][sheet_em][row[COL_BUTTON_ALIAS_EM_IN_EM]] = row

    # Complete buttons and bypass with EM data when possible
    # Description is missing in the main tables but present in the EM tables, so we add it if we can find it via the alias/module
    for b in data["bypass"]:
        module = b.get(COL_BYPASS_NUM_MODULE)
        alias = b.get(COL_BYPASS_ALIAS_EM)
        if module and alias:
            em_data = data["bypass_em"].get(module, {}).get(alias)
            if em_data:
                if em_data.get(COL_BYPASS_DESCRIPTION_ARP):
                    b[COL_BYPASS_DESCRIPTION_ARP] = em_data[COL_BYPASS_DESCRIPTION_ARP]
                if em_data.get(COL_BYPASS_DESCRIPTION_CLIENT):
                    b[COL_BYPASS_DESCRIPTION_CLIENT] = em_data[COL_BYPASS_DESCRIPTION_CLIENT]

    for b in data["buttons"]:
        module = b.get(COL_BUTTON_NUM_MODULE)
        alias = b.get(COL_BUTTON_ALIAS_EM)
        if module and alias:
            em_data = data["buttons_em"].get(module, {}).get(alias)
            if em_data:
                if em_data.get(COL_BUTTON_DESCRIPTION_ARP):
                    b[COL_BUTTON_DESCRIPTION_ARP] = em_data[COL_BUTTON_DESCRIPTION_ARP]
                if em_data.get(COL_BUTTON_DESCRIPTION_CLIENT):
                    b[COL_BUTTON_DESCRIPTION_CLIENT] = em_data[COL_BUTTON_DESCRIPTION_CLIENT]

    return data


# ============================================================
# I/O Console (ask_*)
# ============================================================


def ask_path(prompt: str, allowed_suffixes: Tuple[str, ...]) -> Optional[Path]:
    while True:
        path_str = input(prompt).strip().strip('"')
        if not path_str:
            print("Aucun fichier sélectionné.")
            return None

        try:
            path = Path(path_str).resolve()
        except OSError as e:
            print("Chemin invalide :", e)
            continue

        if not path.exists():
            print("Le fichier n'existe pas.")
            continue

        if not path.is_file():
            print("Le chemin indiqué n'est pas un fichier.")
            continue

        if path.suffix.lower() not in allowed_suffixes:
            print(f"Veuillez sélectionner un fichier parmi : {', '.join(allowed_suffixes)}")
            continue

        return path


def ask_excel_file() -> Optional[Path]:
    return ask_path("Chemin du fichier Excel : ", (".xlsx", ".xlsm", ".xls"))


def ask_bdd_file() -> Optional[Path]:
    return ask_path("Chemin du fichier bdd (sqlite3) : ", (".sqlite3", ".db"))


def ask_input_int(prompt: str) -> int:
    while True:
        try:
            return int(input(prompt).strip())
        except ValueError:
            print("Entrée invalide. Veuillez entrer un nombre.")


def ask_input_str(prompt: str) -> str:
    while True:
        s = input(prompt).strip()
        if s:
            return s
        print("Entrée invalide. Veuillez entrer une chaîne non vide.")


def ask_yes_or_no(prompt: str) -> bool:
    while True:
        response = input(f"{prompt} (o/n) : ").strip().lower()
        if response in ("o", "oui"):
            return True
        if response in ("n", "non"):
            return False
        print("Entrée invalide. Répondez par o/n.")


def ask_language() -> str:
    languages = [LANGUAGE_ARP, LANGUAGE_FR, LANGUAGE_EN, LANGUAGE_ES, LANGUAGE_DE]
    print("Langues disponibles :")
    for i, lang in enumerate(languages):
        print(f"{i} - {lang}")

    while True:
        choice = ask_input_int("Numéro de langue à utiliser pour le client : ")
        if 0 <= choice < len(languages):
            return languages[choice]
        print(f"Veuillez entrer un nombre entre 0 et {len(languages) - 1}.")


# ============================================================
# Exports CSV (textes trad)
# ============================================================


def _csv_line(num_text: int, text: Optional[str]) -> str:
    safe = (text or "").replace('"', '""')  # escape simple CSV-like
    return f'{num_text}:"{safe}";'


def export_defauts_csv(defauts: List[Dict[str, Any]], num_lang: int, out_path: Path) -> None:
    lines: List[str] = []

    for d in defauts:
        code = d.get(COL_DEFAUT_NUM)
        if code is None:
            continue

        try:
            id_ = int(str(code).replace(" ", ""))
        except ValueError:
            print(f"Code défaut invalide : {code}")
            continue

        # ARP
        num_text = (id_ % 1_00_00_00 + BASE_ID_FAULT_DESCRIPTION) * 100
        lines.append(_csv_line(num_text, d.get(COL_DEFAUT_RESOLUTION_ARP)))

        # Client
        num_text = (id_ % 1_00_00_00 + BASE_ID_FAULT_DESCRIPTION) * 100 + num_lang
        lines.append(_csv_line(num_text, d.get(COL_DEFAUT_RESOLUTION_CLIENT)))

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_bypass_csv(bypass_list: List[Dict[str, Any]], num_lang: int, out_path: Path) -> None:
    lines: List[str] = []

    for b in bypass_list:
        num = b.get(COL_BYPASS_NUM)
        if num is None:
            continue

        try:
            id_ = int(num)
        except ValueError:
            print(f"Numéro bypass invalide : {num}")
            continue

        # DESIGNATION ARP / Client
        lines.append(_csv_line((id_ + BASE_ID_BYPASS_TEXT) * 100, b.get(COL_BYPASS_DESIGNATION_ARP)))
        lines.append(_csv_line((id_ + BASE_ID_BYPASS_TEXT) * 100 + num_lang, b.get(COL_BYPASS_DESIGNATION_CLIENT)))

        # DESCRIPTION ARP / Client
        lines.append(_csv_line((id_ + BASE_ID_BYPASS_DESCRIPTION) * 100, b.get(COL_BYPASS_DESCRIPTION_ARP)))
        lines.append(
            _csv_line((id_ + BASE_ID_BYPASS_DESCRIPTION) * 100 + num_lang, b.get(COL_BYPASS_DESCRIPTION_CLIENT))
        )

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def export_button_csv(buttons: List[Dict[str, Any]], num_lang: int, out_path: Path) -> None:
    lines: List[str] = []

    for b in buttons:
        num = b.get(COL_BUTTON_NUM)
        if num is None:
            continue

        try:
            id_ = int(num)
        except ValueError:
            print(f"Numéro bouton invalide : {num}")
            continue

        # DESIGNATION ARP / Client
        lines.append(_csv_line((id_ + BASE_ID_BUTTON_TEXT) * 100, b.get(COL_BUTTON_DESIGNATION_ARP)))
        lines.append(_csv_line((id_ + BASE_ID_BUTTON_TEXT) * 100 + num_lang, b.get(COL_BUTTON_DESIGNATION_CLIENT)))

        # DESCRIPTION ARP / Client
        lines.append(_csv_line((id_ + BASE_ID_BUTTON_DESCRIPTION) * 100, b.get(COL_BUTTON_DESCRIPTION_ARP)))
        lines.append(
            _csv_line((id_ + BASE_ID_BUTTON_DESCRIPTION) * 100 + num_lang, b.get(COL_BUTTON_DESCRIPTION_CLIENT))
        )

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ============================================================
# JSON config (buttons / bypass)
# ============================================================


def ensure_module_cfg(modules_cfg: Dict[str, Dict[str, Any]], module: str) -> Dict[str, Any]:
    """Si le module n'est pas dans modules_cfg, demande à l'utilisateur et l'ajoute."""
    if module not in modules_cfg:
        num_machine = ask_input_int(f"Quel est le numéro de machine pour le module {module} : ")
        num_module = ask_input_int(f"Quel est le numéro de module à utiliser pour le module {module} : ")
        modules_cfg[module] = {"num_machine": num_machine, "num_module": num_module}
    return modules_cfg[module]


def build_buttons_bypass_json(data: Dict[str, Any], num_com: int) -> Dict[str, Any]:
    modules_cfg = data["modules_cfg"]

    json_bypasses = []
    for bypass in data["bypass"]:
        if bypass.get(COL_BYPASS_NUM) is None:
            continue
        module = bypass.get(COL_BYPASS_NUM_MODULE)
        if module is None:
            continue
        if not check_bypass_is_ok(bypass):
            print(f"Le bypass n°{bypass[COL_BYPASS_NUM]} est marqué comme non valide => ignoré.")
            continue

        cfg = ensure_module_cfg(modules_cfg, module)
        json_bypasses.append(
            {
                JSON_BYPASS_NUM: bypass[COL_BYPASS_NUM],
                JSON_BYPASS_NUM_MACHINE: cfg["num_machine"],
                JSON_BYPASS_NUM_MODULE: cfg["num_module"],
                JSON_BYPASS_ALIAS: f"{bypass.get(COL_BYPASS_ALIAS, '')}",
            }
        )

    json_buttons = []
    for button in data["buttons"]:
        if button.get(COL_BUTTON_NUM) is None:
            continue
        module = button.get(COL_BUTTON_NUM_MODULE)
        if module is None:
            continue
        if not check_button_is_ok(button):
            print(f"Le bouton n°{button[COL_BUTTON_NUM]} est marqué comme non valide => ignoré.")
            continue

        cfg = ensure_module_cfg(modules_cfg, module)
        json_buttons.append(
            {
                JSON_BUTTON_NUM: button[COL_BUTTON_NUM],
                JSON_BUTTON_NUM_MACHINE: cfg["num_machine"],
                JSON_BUTTON_NUM_MODULE: cfg["num_module"],
                JSON_BUTTON_ALIAS: f"{button.get(COL_BUTTON_ALIAS, '')}",
            }
        )

    return {"coms": [{"num": num_com, "buttons": json_buttons, "bypasses": json_bypasses}]}


# ============================================================
# SQLite recipes (Format + TradFormat)
# ============================================================

RECIPES_QUERY = """
SELECT 
    f.Numero,
    f.Actif,
    tf.Langue,
    tf.Nom
FROM Format f
LEFT JOIN TRAD_Format tf 
    ON tf.IDFormat = f.IDFormat
ORDER BY f.Numero, tf.Langue;
"""


def fetch_recipes(db_path: Path) -> List[Tuple[Any, Any, Any, Any]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()
        cur.execute(RECIPES_QUERY)
        return cur.fetchall()


def build_recipes(rows: List[Tuple[Any, Any, Any, Any]], num_lang_bdd: int) -> List[Dict[str, Any]]:
    recipes: Dict[int, Dict[str, Any]] = {}

    for numero, actif, langue, nom in rows:
        if numero is None:
            continue

        num = int(numero)
        if num not in recipes:
            recipes[num] = {
                "num": num,
                "name_1": f"Recipe {num}",
                "name_2": f"Recipe {num}",
                "name_3": f"Recipe {num}",
                "used": actif == 1,
                "checked": True,
            }

        if nom is None:
            continue

        if langue == 0:
            recipes[num]["name_1"] = nom
        if langue == num_lang_bdd:
            recipes[num]["name_2"] = nom

    return list(recipes.values())


# ============================================================
# JSON machines
# ============================================================


def capitalize(str):
    return str[:1].upper() + str[1:]


def build_machines(modules_cfg: Dict[str, Dict[str, Any]]) -> Dict[int, Dict[str, Any]]:
    """Regroupe les modules par machine et demande le nom machine une fois."""
    machines: Dict[int, Dict[str, Any]] = {}

    for module, cfg in modules_cfg.items():
        num_machine = cfg["num_machine"]
        if num_machine not in machines:
            machines[num_machine] = {
                "num": num_machine,
                "name_1": capitalize(ask_input_str(f"Nom de la machine n°{num_machine} (langue 1) : ")),
                "name_2": capitalize(ask_input_str(f"Nom de la machine n°{num_machine} (langue 2) : ")),
                "name_3": "",
                "ems": [],
            }

        machines[num_machine]["ems"].append(
            {
                "num": cfg["num_module"],
                "name_1": f"{module} - {capitalize(cfg.get('nom_langue_1', ''))}",
                "name_2": f"{module} - {capitalize(cfg.get('nom_langue_2', ''))}",
                "name_3": f"{module} - {capitalize(cfg.get('nom_langue_3', ''))}",
                "nb_in_machine": cfg["num_module"],
                "utility": 0,
                "checked": True,
                "axs": {},
            }
        )

    return machines


def add_recipes_to_machines(machines: Dict[int, Dict[str, Any]], lang: str) -> None:
    """Pour chaque machine, propose d'ajouter les recipes depuis une DB SQLite."""
    num_lang_bdd = TRANSLATE["bdd"][lang]

    for num_machine, machine in machines.items():
        if not ask_yes_or_no(
            f"Machine {num_machine} - Voulez-vous ajouter les noms des formats ? "
            f"(demandera l'accès à la base de données des recettes)"
        ):
            continue

        db_path = ask_bdd_file()
        if not db_path:
            continue

        rows = fetch_recipes(db_path)
        machine["recipes"] = build_recipes(rows, num_lang_bdd)


# ============================================================
# Main
# ============================================================


def main() -> None:
    excel_path = ask_excel_file()
    if not excel_path:
        return

    print("Lecture du fichier Excel...")
    data = read_excel(excel_path)

    # Langue client + mapping
    lang = ask_language()
    num_lang_defaut = TRANSLATE["defaut"][lang]

    # Exports CSV
    print("Export des CSV...")
    export_defauts_csv(data["defauts"], num_lang_defaut, OUT_DIR / "defaut.csv")
    export_bypass_csv(data["bypass"], num_lang_defaut, OUT_DIR / "bypass.csv")
    export_button_csv(data["buttons"], num_lang_defaut, OUT_DIR / "button.csv")

    # JSON button/bypass
    print("Construction du JSON machines + buttons/bypass...")
    num_com = ask_input_int("Numéro de COM : ")
    j = build_buttons_bypass_json(data, num_com)
    (OUT_DIR / "config_button_bypass.json").write_text(json.dumps(j, ensure_ascii=False, indent=2), encoding="utf-8")

    # JSON machines + recipes
    print("Construction du JSON machines + recipes...")
    machines = build_machines(data["modules_cfg"])
    add_recipes_to_machines(machines, lang)

    out = {"coms": [{"num": num_com, "machines": list(machines.values())}]}
    (OUT_DIR / "config_machines.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
